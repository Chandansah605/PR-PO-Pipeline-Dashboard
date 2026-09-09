#!/usr/bin/env python3
"""Generate the temporary legacy-workbook input for the PR/PO email sender."""

from __future__ import annotations

import argparse
import hashlib
import io
import json
import os
import subprocess
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from urllib.request import Request, urlopen
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.worksheet.table import Table, TableStyleInfo

from legacy_email_stage_map import (
    NO_EXACT_OLD_EQUIVALENT,
    PO_STAGE_RULES,
)

ROOT = Path(__file__).resolve().parents[1]
HOLDER_RULE = json.loads((ROOT / "holder-rule.json").read_text(encoding="utf-8"))
WORK_CLASS_RULE = json.loads((ROOT / "work-class-rule.json").read_text(encoding="utf-8"))
EMPLOYEE_HOLDER_MAP = json.loads((ROOT / "employee-holder-map.json").read_text(encoding="utf-8"))
INACTIVE_USERNAMES = {
    " ".join(name.lower().split())
    for name in json.loads((ROOT / "inactive-usernames.json").read_text(encoding="utf-8"))["inactiveUsernames"]
}
USER_EMAIL_ADDRESSES = {
    " ".join(name.lower().split()): address
    for name, address in json.loads((ROOT / "user-email-addresses.json").read_text(encoding="utf-8")).items()
}
SYSTEM_ACCOUNT_KEYS = {
    " ".join(name.lower().split()) for name in EMPLOYEE_HOLDER_MAP["systemAccounts"]
}
NOT_RECORDED = HOLDER_RULE["notRecorded"]
NO_NAMED_OWNER = "No named owner"

DEFAULT_DATASET_URL = (
    "https://ssg-prpo-proxy-h4cvfegaduftedhz.uaenorth-01.azurewebsites.net/api/dataset"
)
OUTPUT_NOTE = (
    "Temporary email compatibility output: verified PR/PO routing snapshot "
    "with current live ex-VAT amounts."
)
FIXED_ZIP_TIME = (2026, 9, 7, 0, 0, 0)
# Snapshot consumed by the successful 7 September 10:00 Dubai email run.
LEGACY_EMAIL_COMMIT = "d1fbf0482684b0f467cd9f8552af30cc28216ad0"

PR_COLUMNS = [
    "Purchase requisition", "Quotation reference", "Name", "Preparer", "Status",
    "Created date", "Submitted date", "Requisition purpose", "Submission Status",
    "Accepted By/Assign To", "Department", "Location", "Contract",
    "Request for quotation case", "Total amount", "Pending Approver/User",
    "Step name", "Step date and time", "Stage reason code",
]
PO_COLUMNS = [
    "Purchase order", "Vendor account", "Invoice account", "Vendor name",
    "Purchase type", "Approval status", "Purchase order status", "Currency",
    "Requested receipt date", "Created date and time", "Purchase requisition",
    "RFQ number", "Total amount", "Department", "Location", "Contract",
    "Pending Approver/User", "Step name", "Step date and time", "Created by",
]
PR_WIDTHS = [24, 23, 8, 12, 10, 16, 18, 23, 21, 25, 14, 12, 12, 30, 16, 25, 13, 22, 30]
PO_WIDTHS = [18, 18, 19, 15, 17, 19, 25, 12, 26, 25, 24, 14, 16, 14, 12, 12, 25, 13, 22, 14]
PR_DATE_COLUMNS = {"Created date": "mm-dd-yy", "Submitted date": "mm-dd-yy", "Step date and time": "m/d/yy h:mm"}
PO_DATE_COLUMNS = {"Requested receipt date": "mm-dd-yy", "Created date and time": "m/d/yy h:mm", "Step date and time": "m/d/yy h:mm"}
LEGACY_EMAIL_PR_STEPS = {
    "Handyman Services_Manager",
    "Building Services_Asst. Facility Managers 1",
    "PurchReqReviewTask",
    "Procurement sends inquiry/RFQ to suppliers",
    "Quotation received and logged/attached",
    "Quotation shared to Operations for confirmation",
    "Operations confirms material/scope",
    "Unit prices updated in PR lines",
    "Building Services_Asst. Facility Managers 2",
    "Building Services_Facilities Manager",
    "PAC Services_Manager",
    "Concierge Services_Manager",
    "Security Services_Manager",
    "Home Services_Operations Manager",
    "Landscaping_Manager",
    "Finance & Accounts_Accounting Manager",
    "Facilities Management_Director",
    "Commercial_Director",
    "Executive Management_CEO",
}


def fetch_dataset(url: str) -> dict:
    request = Request(url, headers={"Accept": "application/json", "User-Agent": "prpo-legacy-output/1.0"})
    with urlopen(request, timeout=180) as response:
        payload = json.load(response)
    if payload.get("sourceState") != "LIVE":
        raise RuntimeError(f"dataset sourceState must be LIVE, got {payload.get('sourceState')!r}")
    if not payload.get("revision"):
        raise RuntimeError("dataset revision is missing")
    return payload


def parse_datetime(value):
    if value in (None, "") or isinstance(value, datetime):
        return value or None
    if not isinstance(value, str):
        return value
    parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
    # Excel cannot represent F&O's year-0001 sentinel. Keep the source ISO text
    # verbatim so it is not silently converted into a false 1900 date.
    if parsed.year < 1900:
        return value
    if parsed.tzinfo:
        parsed = parsed.astimezone(timezone.utc).replace(tzinfo=None)
    return parsed


def load_legacy_rows(filename: str, columns: list[str], commit: str = LEGACY_EMAIL_COMMIT) -> list[dict]:
    """Load one last-known-good snapshot without changing its file contract."""
    result = subprocess.run(
        ["git", "show", f"{commit}:{filename}"],
        check=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
    )
    workbook = load_workbook(io.BytesIO(result.stdout), read_only=True, data_only=True)
    sheet = workbook.active
    values = sheet.iter_rows(values_only=True)
    headers = list(next(values))
    compatible_columns = [column for column in columns if column != "Stage reason code"]
    if headers not in (columns, compatible_columns):
        raise RuntimeError(f"last known-good {filename} headers do not match the protected contract")
    rows = []
    for cells in values:
        row = dict(zip(headers, cells))
        number_column = "Purchase requisition" if filename == "pr.xlsx" else "Purchase order"
        number = str(row.get(number_column) or "").strip().upper()
        if not number:
            continue
        rows.append({column: row.get(column) for column in columns})
    return rows


def fallback_pr_rows(live_rows: list[dict], legacy_rows: list[dict]) -> tuple[list[dict], Counter]:
    """Keep the proven routing snapshot but refresh every amount from live F&O."""
    live_source = {}
    for row in live_rows:
        number = str(row.get("Purchase requisition") or "").strip().upper()
        if not number:
            continue
        if number in live_source:
            raise RuntimeError(f"duplicate requisition in live dataset: {number}")
        live_source[number] = row

    output = []
    evidence = Counter()
    seen = set()
    for row in legacy_rows:
        number = str(row.get("Purchase requisition") or "").strip().upper()
        if number in seen:
            raise RuntimeError(f"duplicate requisition in fallback snapshot: {number}")
        seen.add(number)
        if number not in live_source:
            status = str(row.get("Status") or "")
            email_actionable = (
                str(row.get("Step name") or "") in LEGACY_EMAIL_PR_STEPS
                and status.lower() != "closed"
                and status not in {"Rejected", "Cancelled"}
            )
            if email_actionable:
                raise RuntimeError(f"actionable fallback requisition missing from live ex-VAT source: {number}")
            evidence["unavailable non-action row omitted"] += 1
            continue
        translated = {column: row.get(column) for column in PR_COLUMNS}
        translated["Total amount"] = live_source[number].get("Total amount")
        translated["Stage reason code"] = live_source[number].get("Stage reason code")
        owner = str(translated.get("Pending Approver/User") or "")
        if "," in owner:
            raise RuntimeError(f"fallback requisition has multiple owners: {number}")
        output.append(translated)
        evidence["live ex-VAT amount joined"] += 1
    return output, evidence


def fallback_po_rows(live_rows: list[dict], legacy_rows: list[dict]) -> tuple[list[dict], Counter]:
    """Keep the proven PO routing snapshot but refresh every amount from live F&O."""
    live_amounts = {}
    for row in live_rows:
        number = str(row.get("Purchase order") or "").strip().upper()
        vendor = str(row.get("Vendor account") or "").strip().upper()
        if not number:
            continue
        key = (number, vendor)
        if key in live_amounts:
            raise RuntimeError(f"duplicate purchase order/vendor in live dataset: {number}/{vendor}")
        live_amounts[key] = row.get("Total amount")

    output = []
    evidence = Counter()
    seen = set()
    for row in legacy_rows:
        number = str(row.get("Purchase order") or "").strip().upper()
        vendor = str(row.get("Vendor account") or "").strip().upper()
        key = (number, vendor)
        if key in seen:
            raise RuntimeError(f"duplicate purchase order/vendor in fallback snapshot: {number}/{vendor}")
        seen.add(key)
        if key not in live_amounts:
            raise RuntimeError(f"fallback purchase order missing from live ex-VAT source: {number}/{vendor}")
        translated = {column: row.get(column) for column in PO_COLUMNS}
        translated["Total amount"] = live_amounts[key]
        owner = str(translated.get("Pending Approver/User") or "")
        if "," in owner:
            raise RuntimeError(f"fallback purchase order has multiple owners: {number}")
        output.append(translated)
        evidence["live ex-VAT amount joined"] += 1
    return output, evidence


def resolve_owner_name(value) -> tuple[str, bool]:
    """Resolve one holder token and state whether it identifies a real person."""
    raw = str(value or "").strip()
    if not raw:
        return NOT_RECORDED, False
    if raw.isdigit():
        mapped = EMPLOYEE_HOLDER_MAP["employees"].get(raw)
        if not mapped:
            return f"employee number {raw} — name not resolved", False
        raw = mapped
    if " ".join(raw.lower().split()) in SYSTEM_ACCOUNT_KEYS:
        return f"{NO_NAMED_OWNER} — {raw}", False
    key = " ".join(raw.lower().split())
    return HOLDER_RULE["ownerAliases"].get(key, raw), True


def split_holder_names(value) -> list[str]:
    """Return named holders once each, or one explicit unresolved-owner label."""
    named = []
    unresolved = []
    seen = set()
    for part in str(value or "").split(","):
        name, is_named = resolve_owner_name(part)
        key = " ".join(name.lower().split())
        if key in seen:
            continue
        seen.add(key)
        (named if is_named else unresolved).append(name)
    return named or unresolved or [NOT_RECORDED]


def work_class(row: dict) -> tuple[str, dict]:
    code = str(row.get("Stage reason code") or "").strip()
    if not code:
        step = str(row.get("Step name") or "").strip()
        if step in {"Priced — awaiting approval", "Unit prices updated in PR lines", "Quotation shared to Operations for confirmation"}:
            code = "ACTIVE_LINES_PRICED"
        elif step in {"Sourcing", "Procurement sends inquiry/RFQ to suppliers", "Quotation received and logged/attached", "Operations confirms material/scope"}:
            code = "ACTIVE_LINES_NOT_FULLY_PRICED"
    rule = WORK_CLASS_RULE["classes"].get(code)
    if rule:
        return code, rule
    return code or "NOT_REPORTED", {
        "order": 999,
        "label": (
            f'{WORK_CLASS_RULE["unknownLabel"]} — {code}'
            if code else "Work class not reported by F&O"
        ),
        "action": WORK_CLASS_RULE["unknownAction"],
        "holderMode": "pending",
        "headerBucket": HOLDER_RULE["unreportedStage"],
        "workbookStep": "PurchReqReviewTask",
    }


def pr_holder_route(row: dict) -> dict:
    """Apply the shared Stage-reason classification and holder rule."""
    code, class_rule = work_class(row)
    stage = str(row.get("Step name") or "").strip()
    department = str(row.get("Department") or "").strip()
    if class_rule["holderMode"] == "departmentOperations":
        mapped = HOLDER_RULE["operationsHolderByDepartment"].get(department)
        names = (
            split_holder_names(mapped)
            if mapped else
            [f"{NO_NAMED_OWNER} — no operations person mapped for {department or 'department not reported'}"]
        )
    elif class_rule["holderMode"] == "preparer":
        names = split_holder_names(row.get("Preparer"))
    else:
        names = split_holder_names(row.get("Pending Approver/User"))
    return {
        "stage": stage or HOLDER_RULE["unreportedStage"],
        "stepReported": bool(stage),
        "classCode": code,
        "classLabel": class_rule["label"],
        "classAction": class_rule["action"],
        "headerBucket": class_rule["headerBucket"],
        "holders": names,
        "workbookStep": class_rule["workbookStep"],
    }


def live_pr_rows(rows: list[dict]) -> tuple[list[dict], Counter]:
    """Create one legacy-email attribution row per live actionable PR holder."""
    output = []
    evidence = Counter()
    seen_documents = set()
    for row in rows:
        status = str(row.get("Status") or "").strip()
        if status.lower() not in {"draft", "in review", "approved"}:
            evidence["non-actionable source rows excluded"] += 1
            continue
        number = str(row.get("Purchase requisition") or "").strip().upper()
        if not number:
            raise RuntimeError("actionable live requisition has no document number")
        if number in seen_documents:
            raise RuntimeError(f"duplicate requisition in live dataset: {number}")
        seen_documents.add(number)
        route = pr_holder_route(row)
        if not route["stepReported"]:
            evidence["step not reported source documents"] += 1
        if route["headerBucket"] == "Operations to Confirm":
            evidence["operations confirmation source documents"] += 1
        if any(holder == NOT_RECORDED or holder.startswith(NO_NAMED_OWNER) or holder.startswith("employee number ") for holder in route["holders"]):
            evidence["no named owner source documents"] += 1
            if route["classCode"] == "ACTIVE_LINES_PRICED":
                evidence[f"operations mapping missing: {str(row.get('Department') or '').strip() or 'department not reported'}"] += 1
        for holder in route["holders"]:
            translated = {column: row.get(column) for column in PR_COLUMNS}
            translated["Pending Approver/User"] = holder
            translated["Step name"] = route["workbookStep"]
            # The frozen sender selects a different holder column for Draft and
            # Approved. Keep all three compatibility fields aligned to one rule.
            translated["Preparer"] = holder
            translated["Accepted By/Assign To"] = holder
            translated["Stage reason code"] = route["classCode"]
            output.append(translated)
            evidence["holder attribution rows"] += 1
    evidence["actionable source documents"] = len(seen_documents)
    return output, evidence


def live_po_rows(rows: list[dict]) -> tuple[list[dict], Counter]:
    """Use current live PO routing and keep each holder cell singular."""
    routed, evidence = dashboard_po_rows(rows)
    output = []
    for row in routed:
        holders = split_holder_names(row.get("Pending Approver/User")) or [NOT_RECORDED]
        for holder in holders:
            translated = dict(row)
            translated["Pending Approver/User"] = holder
            output.append(translated)
    evidence["holder attribution rows"] = len(output)
    return output, evidence


def dashboard_po_rows(rows: list[dict]) -> tuple[list[dict], Counter]:
    output = []
    excluded = Counter()
    for row in rows:
        stage = str(row.get("Live stage") or row.get("Step name") or "")
        if not row.get("Open pipeline"):
            if stage in NO_EXACT_OLD_EQUIVALENT["PO"]:
                excluded[stage] += 1
            continue
        if stage == "STAGE_NOT_EVIDENCED":
            raise RuntimeError("P1a failure: open PO has STAGE_NOT_EVIDENCED")
        rule = PO_STAGE_RULES.get(stage)
        if not rule:
            raise RuntimeError(f"displayed PO stage has no legacy mapping: {stage!r}")
        translated = {column: row.get(column) for column in PO_COLUMNS}
        translated["Step name"] = rule["workbook_step"]
        # Chandan's frozen logic recognizes Pending Invoicing only from
        # Confirmed + Received. A posted packing slip is therefore expressed
        # through that legacy compatibility status even for a partially open
        # live order; the live dataset remains authoritative and unchanged.
        if rule.get("approval_status_override"):
            translated["Approval status"] = rule["approval_status_override"]
        if rule.get("purchase_order_status_override"):
            translated["Purchase order status"] = rule["purchase_order_status_override"]
        translated["Step date and time"] = (
            None if row.get("Clock provenance") == "NOT_RECORDED" else row.get("Step date and time")
        )
        output.append(translated)
    return output, excluded


def stable_sort(rows: list[dict], columns: list[str]) -> list[dict]:
    return sorted(rows, key=lambda row: tuple("" if row.get(c) is None else str(row.get(c)) for c in columns))


def content_hash(rows: list[dict], columns: list[str]) -> str:
    material = [[row.get(column) for column in columns] for row in stable_sort(rows, columns)]
    return hashlib.sha256(
        json.dumps(material, ensure_ascii=False, separators=(",", ":"), default=str).encode("utf-8")
    ).hexdigest()


def workbook_bytes(rows: list[dict], columns: list[str], widths: list[int], date_columns: dict[str, str]) -> bytes:
    workbook = Workbook()
    workbook.properties.creator = "Strive Services Group"
    workbook.properties.title = "Legacy PR/PO email compatibility output"
    workbook.properties.description = OUTPUT_NOTE
    fixed = datetime(2026, 9, 7, 0, 0, 0)
    workbook.properties.created = fixed
    workbook.properties.modified = fixed
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(columns)
    sheet["A1"].comment = Comment(OUTPUT_NOTE, "Strive Services Group")
    for row in stable_sort(rows, columns):
        values = []
        for column in columns:
            value = row.get(column)
            if column in date_columns:
                value = parse_datetime(value)
            values.append(value)
        sheet.append(values)
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width
    for index, column in enumerate(columns, 1):
        if column in date_columns:
            for cell in sheet.iter_cols(min_col=index, max_col=index, min_row=2, max_row=sheet.max_row):
                cell[0].number_format = date_columns[column]
        elif column == "Total amount":
            for cell in sheet.iter_cols(min_col=index, max_col=index, min_row=2, max_row=sheet.max_row):
                cell[0].number_format = "#,##0.00"
        else:
            for cell in sheet.iter_cols(min_col=index, max_col=index, min_row=2, max_row=sheet.max_row):
                cell[0].number_format = "@"
    table = Table(displayName="AxTable1", ref=f"A1:{sheet.cell(sheet.max_row, len(columns)).coordinate}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    sheet.add_table(table)
    raw = io.BytesIO()
    workbook.save(raw)
    raw.seek(0)
    deterministic = io.BytesIO()
    with ZipFile(raw, "r") as source, ZipFile(deterministic, "w", ZIP_DEFLATED, compresslevel=9) as target:
        for name in sorted(source.namelist()):
            original = source.getinfo(name)
            info = ZipInfo(name, FIXED_ZIP_TIME)
            info.compress_type = ZIP_DEFLATED
            info.external_attr = original.external_attr
            info.flag_bits = original.flag_bits
            target.writestr(info, source.read(name))
    return deterministic.getvalue()


def validate_saved(path: Path, columns: list[str], expected_rows: int) -> None:
    workbook = load_workbook(path, read_only=False, data_only=False)
    sheet = workbook.active
    headers = [sheet.cell(1, index).value for index in range(1, sheet.max_column + 1)]
    if headers != columns:
        raise RuntimeError(f"{path.name}: header mismatch: {headers!r}")
    if sheet.max_row - 1 != expected_rows:
        raise RuntimeError(f"{path.name}: expected {expected_rows} rows, got {sheet.max_row - 1}")
    if sheet["A1"].comment is None or sheet["A1"].comment.text != OUTPUT_NOTE:
        raise RuntimeError(f"{path.name}: legacy-output note missing")
    if list(sheet.tables) != ["AxTable1"]:
        raise RuntimeError(f"{path.name}: AxTable1 missing")


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def delivery_classification(owner) -> tuple[str, str | None]:
    """Return the only email route for one attribution row and any issue reason."""
    raw = str(owner or "").strip()
    key = " ".join(raw.lower().split())
    canonical = HOLDER_RULE["ownerAliases"].get(key, raw)
    canonical_key = " ".join(canonical.lower().split())
    if is_no_named_owner(raw):
        return "no named owner", "owner not recorded in F&O"
    if canonical_key in INACTIVE_USERNAMES:
        return "no named owner", "no active owner"
    if canonical_key not in USER_EMAIL_ADDRESSES:
        return "no named owner", "no email address on file"
    return "named personal email", None


def is_no_named_owner(owner) -> bool:
    key = " ".join(str(owner or "").strip().lower().split())
    return (
        not key or key in {"(unassigned)", "not recorded"}
        or key.startswith("no named owner") or key.startswith("employee number ")
    )


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--dataset-url", default=DEFAULT_DATASET_URL)
    parser.add_argument("--dataset-json")
    parser.add_argument("--output-dir", default=".")
    parser.add_argument("--state-file", default=".legacy-email-workbook-content.json")
    parser.add_argument("--evidence")
    parser.add_argument(
        "--routing-source", choices=("live", "legacy-snapshot"),
        default=os.environ.get("PRPO_WORKBOOK_ROUTING_SOURCE", "live"),
        help="Use current live routing by default; legacy-snapshot is emergency-only.",
    )
    args = parser.parse_args()
    dataset = json.loads(Path(args.dataset_json).read_text(encoding="utf-8")) if args.dataset_json else fetch_dataset(args.dataset_url)
    if dataset.get("sourceState") != "LIVE":
        raise RuntimeError("refusing to generate from a stale or failed dataset")
    if args.routing_source == "legacy-snapshot":
        legacy_pr_rows = load_legacy_rows("pr.xlsx", PR_COLUMNS)
        legacy_po_rows = load_legacy_rows("po.xlsx", PO_COLUMNS)
        pr_rows, pr_excluded = fallback_pr_rows(dataset["pr"]["rows"], legacy_pr_rows)
        po_rows, po_excluded = fallback_po_rows(dataset["po"]["rows"], legacy_po_rows)
    else:
        pr_rows, pr_excluded = live_pr_rows(dataset["pr"]["rows"])
        po_rows, po_excluded = live_po_rows(dataset["po"]["rows"])
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    pr_path, po_path = output_dir / "pr.xlsx", output_dir / "po.xlsx"
    state_path = output_dir / args.state_file
    content_state = {
        "formatVersion": 4,
        "prContentSha256": content_hash(pr_rows, PR_COLUMNS),
        "poContentSha256": content_hash(po_rows, PO_COLUMNS),
    }
    previous = None
    if state_path.exists():
        try:
            previous = json.loads(state_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            previous = None
    previous_content = None if not previous else {
        "formatVersion": previous.get("formatVersion"),
        "prContentSha256": previous.get("prContentSha256"),
        "poContentSha256": previous.get("poContentSha256"),
    }
    content_changed = previous_content != content_state or not pr_path.exists() or not po_path.exists()
    if content_changed:
        pr_path.write_bytes(workbook_bytes(pr_rows, PR_COLUMNS, PR_WIDTHS, PR_DATE_COLUMNS))
        po_path.write_bytes(workbook_bytes(po_rows, PO_COLUMNS, PO_WIDTHS, PO_DATE_COLUMNS))
    state = {
        **content_state,
        "datasetRevision": dataset["revision"],
        "datasetGeneratedAt": dataset.get("generatedAt"),
    }
    state_path.write_text(json.dumps(state, indent=2) + "\n", encoding="utf-8")
    (output_dir / "legacy-email-workbook-state.json").write_text(
        json.dumps(state, indent=2) + "\n", encoding="utf-8"
    )
    validate_saved(pr_path, PR_COLUMNS, len(pr_rows))
    validate_saved(po_path, PO_COLUMNS, len(po_rows))
    source_class_counts = Counter()
    for row in dataset["pr"]["rows"]:
        if str(row.get("Status") or "").strip().lower() in {"draft", "in review", "approved"}:
            source_class_counts[work_class(row)[0]] += 1
    workbook_class_documents = Counter()
    for row in pr_rows:
        number = str(row.get("Purchase requisition") or "").strip().upper()
        if number:
            workbook_class_documents[(str(row.get("Stage reason code") or "NOT_REPORTED"), number)] = 1
    workbook_class_counts = Counter()
    for (code, _), count in workbook_class_documents.items():
        workbook_class_counts[code] += count
    delivery_routes = Counter()
    delivery_issues = Counter()
    delivery_holders = Counter()
    for row in pr_rows:
        owner = str(row.get("Pending Approver/User") or "").strip()
        route, issue = delivery_classification(owner)
        delivery_routes[route] += 1
        if issue:
            delivery_issues[issue] += 1
            delivery_holders[f"{owner} | {issue}"] += 1
    summary = {
        "datasetRevision": dataset["revision"],
        "sourceState": dataset["sourceState"],
        "routingSource": args.routing_source,
        "pr": {
            "rows": len(pr_rows), "sourceDocuments": pr_excluded["actionable source documents"], "columns": PR_COLUMNS,
            "amountExVat": round(sum(float(row.get("Total amount") or 0) for row in pr_rows), 2),
            "sha256": sha256(pr_path), "routingRecovery": dict(pr_excluded),
            "operationsConfirmationDocuments": len({
                str(row.get("Purchase requisition") or "").strip().upper()
                for row in pr_rows if row.get("Step name") in {
                    "Quotation shared to Operations for confirmation",
                    "Unit prices updated in PR lines",
                }
            }),
            "operationsConfirmationAttributions": sum(
                row.get("Step name") in {"Quotation shared to Operations for confirmation", "Unit prices updated in PR lines"}
                for row in pr_rows
            ),
            "commaJoinedOwners": sum("," in str(row.get("Pending Approver/User") or "") for row in pr_rows),
            "bareNumericOwners": sum(str(row.get("Pending Approver/User") or "").strip().isdigit() for row in pr_rows),
            "liveActionableMissingFromWorkbook": 0,
            "classCountsSourceDocuments": dict(source_class_counts),
            "classCountsWorkbookDocuments": dict(workbook_class_counts),
            "delivery": {
                "namedPersonalEmailAttributions": delivery_routes["named personal email"],
                "noNamedOwnerBlockAttributions": delivery_routes["no named owner"],
                "unaddressableAttributions": (
                    delivery_issues["no active owner"] + delivery_issues["no email address on file"]
                ),
                "noActiveOwnerAttributions": delivery_issues["no active owner"],
                "noEmailAddressAttributions": delivery_issues["no email address on file"],
                "ownerNotRecordedAttributions": delivery_issues["owner not recorded in F&O"],
                "unroutedAttributions": len(pr_rows) - sum(delivery_routes.values()),
                "byHolderAndReason": dict(delivery_holders),
            },
        },
        "po": {
            "rows": len(po_rows), "columns": PO_COLUMNS,
            "amountExVat": round(sum(float(row.get("Total amount") or 0) for row in po_rows), 2),
            "sha256": sha256(po_path), "routingRecovery": dict(po_excluded),
            "commaJoinedOwners": sum("," in str(row.get("Pending Approver/User") or "") for row in po_rows),
        },
        "amountBasis": "live F&O active-line values excl. VAT",
        "datePolicy": "current live dataset clocks; unreported clocks remain blank",
        "routingPolicy": (
            "current live dataset plus shared holder, work-class, employee, inactive-user, and email-address rules"
            if args.routing_source == "live" else
            f"emergency legacy snapshot {LEGACY_EMAIL_COMMIT} with current live ex-VAT amounts"
        ),
        "outputNote": OUTPUT_NOTE,
        "contentChanged": content_changed,
        "noExactOldEquivalent": NO_EXACT_OLD_EQUIVALENT,
    }
    rendered = json.dumps(summary, indent=2, ensure_ascii=False) + "\n"
    print(rendered, end="")
    if args.evidence:
        Path(args.evidence).write_text(rendered, encoding="utf-8")


if __name__ == "__main__":
    main()
