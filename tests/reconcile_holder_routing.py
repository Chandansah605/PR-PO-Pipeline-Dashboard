import argparse
import json
import subprocess
import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))
import generate_legacy_email_workbooks as generator


def workbook_rows(path: Path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    values = workbook.active.iter_rows(values_only=True)
    headers = list(next(values))
    return headers, [dict(zip(headers, cells)) for cells in values]


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("dataset")
    parser.add_argument("pr_workbook")
    parser.add_argument("po_workbook")
    args = parser.parse_args()

    dataset = json.loads(Path(args.dataset).read_text(encoding="utf-8"))
    dashboard = json.loads(subprocess.check_output(
        ["node", "tests/dashboard-holder-snapshot.js", args.dataset], cwd=ROOT, text=True
    ))
    pr_headers, pr_rows = workbook_rows(Path(args.pr_workbook))
    po_headers, po_rows = workbook_rows(Path(args.po_workbook))
    assert pr_headers == generator.PR_COLUMNS
    assert po_headers == generator.PO_COLUMNS

    actionable = [row for row in dataset["pr"]["rows"] if str(row.get("Status") or "").strip().lower() in {"draft", "in review", "approved"}]
    source_numbers = {str(row["Purchase requisition"]).strip().upper() for row in actionable}
    workbook_numbers = {str(row["Purchase requisition"]).strip().upper() for row in pr_rows}
    missing = sorted(source_numbers - workbook_numbers)
    assert not missing, f"live actionable PRs missing from workbook: {missing[:10]}"
    assert not any("," in str(row.get("Pending Approver/User") or "") for row in pr_rows + po_rows)

    buyers = ["adnan.ullah", "aparna.pauly", "layusha.cleatus", "roderick.red"]
    email_counts = Counter(str(row.get("Pending Approver/User") or "").strip().lower() for row in pr_rows)
    email_buyers = {buyer: email_counts[buyer] for buyer in buyers}
    assert email_buyers == dashboard["holderCounts"], (email_buyers, dashboard["holderCounts"])

    ops_rows = [row for row in pr_rows if row.get("Step name") == "Unit prices updated in PR lines"]
    ops_documents = {str(row["Purchase requisition"]).strip().upper() for row in ops_rows}
    assert len(ops_documents) == dashboard["operationsConfirmation"]
    assert dashboard["gates"].get("Prices Updated") == dashboard["operationsConfirmation"]
    assert dashboard["commaJoinedLabels"] == 0

    summary = {
        "datasetRevision": dataset["revision"],
        "liveActionableDocuments": len(source_numbers),
        "workbookAttributionRows": len(pr_rows),
        "liveActionableMissingFromWorkbook": len(missing),
        "dashboardUniqueAmountExVat": dashboard["amountExVat"],
        "dashboardStepNotReported": dashboard["stepNotReported"],
        "operationsConfirmationDashboard": dashboard["operationsConfirmation"],
        "operationsConfirmationEmail": len(ops_documents),
        "operationsConfirmationAttributions": len(ops_rows),
        "pricesUpdatedJourneyGate": dashboard["gates"].get("Prices Updated", 0),
        "buyerCounts": {buyer: {"dashboard": dashboard["holderCounts"][buyer], "email": email_buyers[buyer]} for buyer in buyers},
        "prHeadersUnchanged": pr_headers == generator.PR_COLUMNS,
        "poHeadersUnchanged": po_headers == generator.PO_COLUMNS,
        "commaJoinedHolderCells": 0,
    }
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
