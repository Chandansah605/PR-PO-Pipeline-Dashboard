import io
import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

import generate_legacy_email_workbooks as generator


def legacy_row(**overrides):
    row = {column: None for column in generator.PR_COLUMNS}
    row.update({
        "Purchase requisition": "PR-TEST-001",
        "Status": "In review",
        "Step name": "Unit prices updated in PR lines",
        "Pending Approver/User": "procurement.user",
        "Step date and time": "2026-09-04T06:00:00Z",
        "Total amount": 105.0,
    })
    row.update(overrides)
    return row


def legacy_po_row(**overrides):
    row = {column: None for column in generator.PO_COLUMNS}
    row.update({
        "Purchase order": "PO-TEST-001",
        "Vendor account": "VEND-001",
        "Approval status": "In review",
        "Purchase order status": "Backorder",
        "Step name": "Accounting Manager",
        "Pending Approver/User": "finance.user",
        "Step date and time": "2026-09-04T06:00:00Z",
        "Total amount": 105.0,
    })
    row.update(overrides)
    return row


class LegacyEmailFallbackTests(unittest.TestCase):
    def test_default_main_path_never_reads_the_frozen_snapshot(self):
        dataset = {
            "revision": "test-live-revision", "sourceState": "LIVE",
            "pr": {"rows": [legacy_row(**{"Step name": "Sourcing"})]},
            "po": {"rows": [legacy_po_row(**{"Live stage": "Not yet sent", "Open pipeline": True})]},
        }
        with tempfile.TemporaryDirectory() as output_dir, patch.object(
            generator, "fetch_dataset", return_value=dataset
        ), patch.object(
            generator, "load_legacy_rows", side_effect=AssertionError("snapshot read")
        ), patch.object(sys, "argv", ["generate", "--output-dir", output_dir]):
            generator.main()

    def test_live_holder_split_deduplicates_case_insensitively(self):
        source = legacy_row(**{
            "Step name": "Sourcing",
            "Pending Approver/User": "Adnan.Ullah, adnan.ullah, Layusha.cleatus",
        })
        rows, evidence = generator.live_pr_rows([source])
        self.assertEqual([row["Pending Approver/User"] for row in rows], ["Adnan.Ullah", "Layusha.cleatus"])
        self.assertEqual(evidence["actionable source documents"], 1)

    def test_live_single_holder_stays_single(self):
        rows, _ = generator.live_pr_rows([legacy_row(**{
            "Step name": "Sourcing", "Pending Approver/User": "Aparna.Pauly"
        })])
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["Pending Approver/User"], "Aparna.Pauly")

    def test_live_blank_holder_is_explicit(self):
        rows, _ = generator.live_pr_rows([legacy_row(**{
            "Step name": "Sourcing", "Pending Approver/User": ""
        })])
        self.assertEqual(rows[0]["Pending Approver/User"], "not recorded")

    def test_live_missing_step_is_kept_with_compatibility_route(self):
        rows, evidence = generator.live_pr_rows([legacy_row(**{
            "Purchase requisition": "PR-NEW-AFTER-SNAPSHOT",
            "Step name": None,
            "Stage reason code": "UNMAPPED_ELEMENT",
            "Pending Approver/User": "roderick.red",
        })])
        self.assertEqual(rows[0]["Purchase requisition"], "PR-NEW-AFTER-SNAPSHOT")
        self.assertEqual(rows[0]["Step name"], "PurchReqReviewTask")
        self.assertEqual(evidence["step not reported source documents"], 1)

    def test_priced_routes_to_department_operations_confirmer(self):
        rows, evidence = generator.live_pr_rows([legacy_row(**{
            "Step name": "Priced — awaiting approval",
            "Stage reason code": "ACTIVE_LINES_PRICED",
            "Department": "Building Services",
            "Pending Approver/User": "Adnan.Ullah",
        })])
        self.assertEqual(rows[0]["Pending Approver/User"], "dinesh.laxman")
        self.assertEqual(rows[0]["Step name"], "Unit prices updated in PR lines")
        self.assertEqual(evidence["operations confirmation source documents"], 1)

    def test_priced_without_department_mapping_goes_to_no_named_owner(self):
        rows, evidence = generator.live_pr_rows([legacy_row(**{
            "Step name": "Priced — awaiting approval",
            "Stage reason code": "ACTIVE_LINES_PRICED",
            "Department": "Surveying Services",
            "Pending Approver/User": "Aparna.Pauly",
        })])
        self.assertEqual(rows[0]["Pending Approver/User"], "No named owner — no operations person mapped for Surveying Services")
        self.assertEqual(evidence["operations mapping missing: Surveying Services"], 1)

    def test_every_stage_reason_code_has_a_plain_english_class(self):
        for code, rule in generator.WORK_CLASS_RULE["classes"].items():
            with self.subTest(code=code):
                _, resolved = generator.work_class(legacy_row(**{"Stage reason code": code}))
                self.assertEqual(resolved["label"], rule["label"])
                self.assertTrue(resolved["action"])

    def test_preparer_employee_number_resolves_to_a_name(self):
        rows, _ = generator.live_pr_rows([legacy_row(**{
            "Step name": None,
            "Stage reason code": "NO_CURRENT_WORK_ITEM",
            "Preparer": "310523",
            "Pending Approver/User": None,
        })])
        self.assertEqual(rows[0]["Pending Approver/User"], "dinesh.laxman")
        self.assertEqual(rows[0]["Preparer"], "dinesh.laxman")

    def test_unknown_preparer_employee_number_is_explicit(self):
        rows, evidence = generator.live_pr_rows([legacy_row(**{
            "Step name": None,
            "Stage reason code": "NO_CURRENT_WORK_ITEM",
            "Preparer": "999999",
            "Pending Approver/User": None,
        })])
        self.assertEqual(rows[0]["Pending Approver/User"], "employee number 999999 — name not resolved")
        self.assertEqual(evidence["no named owner source documents"], 1)

    def test_system_account_preparer_goes_to_no_named_owner(self):
        rows, evidence = generator.live_pr_rows([legacy_row(**{
            "Step name": None,
            "Stage reason code": "NO_CURRENT_WORK_ITEM",
            "Preparer": "000000",
            "Pending Approver/User": None,
        })])
        self.assertEqual(rows[0]["Pending Approver/User"], "No named owner — D365CRM ADMIN")
        self.assertEqual(evidence["no named owner source documents"], 1)

    def test_delivery_classifies_inactive_unaddressed_and_addressed_holders(self):
        self.assertEqual(
            generator.delivery_classification("Layusha.cleatus"),
            ("no named owner", "no active owner"),
        )
        self.assertEqual(
            generator.delivery_classification("Sirinikhil"),
            ("no named owner", "no email address on file"),
        )
        self.assertEqual(
            generator.delivery_classification("Zaheer Ahmed Ameer"),
            ("named personal email", None),
        )

    def test_delivery_classification_covers_every_attribution_once(self):
        owners = ["Layusha.cleatus", "Sirinikhil", "Zaheer.Ahmed", "No named owner — IT DEPARTMENT"]
        routes = [generator.delivery_classification(owner)[0] for owner in owners]
        self.assertEqual(routes.count("named personal email"), 1)
        self.assertEqual(routes.count("no named owner"), 3)
        self.assertEqual(len(routes), 4)

    def test_preserves_routing_and_replaces_amount_from_live_source(self):
        live = [{"Purchase requisition": "pr-test-001", "Total amount": 100.0}]
        rows, evidence = generator.fallback_pr_rows(live, [legacy_row()])
        self.assertEqual(rows[0]["Step name"], "Unit prices updated in PR lines")
        self.assertEqual(rows[0]["Pending Approver/User"], "procurement.user")
        self.assertEqual(rows[0]["Step date and time"], "2026-09-04T06:00:00Z")
        self.assertEqual(rows[0]["Total amount"], 100.0)
        self.assertEqual(evidence["live ex-VAT amount joined"], 1)

    def test_fails_if_an_actionable_snapshot_row_has_no_live_amount(self):
        with self.assertRaisesRegex(RuntimeError, "actionable fallback requisition"):
            generator.fallback_pr_rows([], [legacy_row()])

    def test_omits_only_a_non_actionable_row_missing_from_live_source(self):
        row = legacy_row(Status="Closed")
        rows, evidence = generator.fallback_pr_rows([], [row])
        self.assertEqual(rows, [])
        self.assertEqual(evidence["unavailable non-action row omitted"], 1)

    def test_rejects_a_comma_joined_owner(self):
        live = [{"Purchase requisition": "PR-TEST-001", "Total amount": 100.0}]
        row = legacy_row(**{"Pending Approver/User": "one.user, two.user"})
        with self.assertRaisesRegex(RuntimeError, "multiple owners"):
            generator.fallback_pr_rows(live, [row])

    def test_generated_workbook_keeps_exact_header_contract(self):
        metadata = [
            {"Purchase requisition": "PR-TEST-001", "Source holder": "one.user"},
            {"Purchase requisition": "PR-TEST-001", "Source holder": "two.user"},
        ]
        payload = generator.workbook_bytes(
            [legacy_row()], generator.PR_COLUMNS, generator.PR_WIDTHS,
            generator.PR_DATE_COLUMNS, metadata
        )
        workbook = load_workbook(io.BytesIO(payload), read_only=False, data_only=True)
        sheet = workbook.active
        headers = [sheet.cell(1, index).value for index in range(1, sheet.max_column + 1)]
        self.assertEqual(headers, generator.PR_COLUMNS)
        self.assertEqual(headers[-1], "Stage reason code")
        self.assertEqual(list(sheet.tables), ["AxTable1"])
        self.assertEqual(workbook["Routing metadata"].sheet_state, "hidden")
        self.assertEqual(list(workbook["Routing metadata"].values)[1:], [
            ("PR-TEST-001", "one.user"), ("PR-TEST-001", "two.user")
        ])

    def test_shared_metadata_keeps_one_holder_per_row(self):
        row = legacy_row(**{
            "Stage reason code": "ACTIVE_LINES_NOT_FULLY_PRICED",
            "Pending Approver/User": "Adnan.Ullah, Adnan.Ullah, Layusha.cleatus, roderick.red",
        })
        metadata = generator.shared_routing_metadata([row])
        self.assertEqual(metadata, [
            {"Purchase requisition": "PR-TEST-001", "Source holder": "Adnan.Ullah"},
            {"Purchase requisition": "PR-TEST-001", "Source holder": "Layusha.cleatus"},
            {"Purchase requisition": "PR-TEST-001", "Source holder": "roderick.red"},
        ])
        self.assertFalse(any("," in item["Source holder"] for item in metadata))

    def test_po_fallback_preserves_routing_and_replaces_live_amount(self):
        live = [{"Purchase order": "po-test-001", "Vendor account": "vend-001", "Total amount": 100.0}]
        rows, evidence = generator.fallback_po_rows(live, [legacy_po_row()])
        self.assertEqual(rows[0]["Step name"], "Accounting Manager")
        self.assertEqual(rows[0]["Pending Approver/User"], "finance.user")
        self.assertEqual(rows[0]["Total amount"], 100.0)
        self.assertEqual(evidence["live ex-VAT amount joined"], 1)

    def test_po_fallback_fails_without_a_live_amount(self):
        with self.assertRaisesRegex(RuntimeError, "purchase order missing"):
            generator.fallback_po_rows([], [legacy_po_row()])

    def test_po_amount_join_disambiguates_reused_numbers_by_vendor(self):
        live = [
            {"Purchase order": "PO-TEST-001", "Vendor account": "VEND-001", "Total amount": 100.0},
            {"Purchase order": "PO-TEST-001", "Vendor account": "VEND-002", "Total amount": 900.0},
        ]
        rows, _ = generator.fallback_po_rows(live, [legacy_po_row()])
        self.assertEqual(rows[0]["Total amount"], 100.0)


if __name__ == "__main__":
    unittest.main()
